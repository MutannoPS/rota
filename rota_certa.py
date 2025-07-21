from openpyxl import load_workbook, Workbook
import requests, re, io, os, asyncio
from datetime import datetime
from flask import Flask, request
from telegram import Update
from telegram.ext import Application, MessageHandler, ContextTypes, filters

TOKEN = os.environ.get("TOKEN", "8095673432:AAGa19vnVQDqxLDz_OSr0wFPQUzH2mh03sA")

# Flask e Telegram apps
app_web = Flask(__name__)
app_telegram = Application.builder().token(TOKEN).build()

def consultar_logradouro(cep):
    cep = str(cep).replace("-", "").strip()
    url = f"https://opencep.com/v1/{cep}"
    try:
        response = requests.get(url)
        return response.json().get("logradouro") if response.status_code == 200 else None
    except:
        return None

def extrair_numero_complemento(endereco):
    match = re.search(r"\d.*", endereco)
    return match.group(0) if match else ""

def corrigir_endereco(endereco, logradouro_api):
    numero = extrair_numero_complemento(endereco)
    return f"{logradouro_api}, {numero}".strip(", ") if logradouro_api else endereco

def normalizar_endereco(endereco):
    match = re.match(r"(.+?),\s*(\d+)", endereco)
    return f"{match.group(1).strip()}, {match.group(2).strip()}" if match else endereco.strip()

def formatar_sequence(lista):
    itens = sorted(set(lista))
    texto = ", ".join(itens)
    return f"{texto}; Total: {len(itens)} pacotes." if len(itens) > 1 else texto

async def tratar_arquivo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    hora = datetime.now().hour
    saudacao = "Bom dia ☀️" if hora < 12 else "Boa tarde 🌤️" if hora < 18 else "Boa noite 🌙"
    await update.message.reply_text(f"{saudacao}\n\n📥 Aprimorando sua planilha... ⏳")

    arquivo = update.message.document
    nome_original = arquivo.file_name
    match = re.search(r"(\d{2}-\d{2}-\d{4})", nome_original)
    data_str = match.group(1) if match else "data"
    nome_final = f"RotaAtualizada-{data_str}.xlsx"

    file_bytes = await (await arquivo.get_file()).download_as_bytearray()
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    dados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        linha = dict(zip(headers, row))
        cep = linha.get("Zipcode/Postal code")
        endereco = linha.get("Destination Address")
        if cep and endereco:
            linha["Destination Address"] = corrigir_endereco(endereco, consultar_logradouro(cep))
        linha["Endereco Corrigido"] = normalizar_endereco(linha.get("Destination Address", ""))
        linha["Sequence"] = str(linha.get("Sequence", ""))
        dados.append(linha)

    agrupado = {}
    for item in dados:
        chave = item["Endereco Corrigido"]
        if chave not in agrupado:
            agrupado[chave] = {
                "Destination Address": item["Destination Address"],
                "Bairro": item.get("Bairro", ""),
                "City": item.get("City", ""),
                "Zipcode/Postal code": item.get("Zipcode/Postal code", ""),
                "Sequence": []
            }
        agrupado[chave]["Sequence"].append(item["Sequence"])

    wb_saida = Workbook()
    ws_saida = wb_saida.active
    ws_saida.append(["Sequence", "Destination Address", "Bairro", "City", "Zipcode/Postal code"])

    for valor in agrupado.values():
        ws_saida.append([
            formatar_sequence(valor["Sequence"]),
            valor["Destination Address"],
            valor["Bairro"],
            valor["City"],
            valor["Zipcode/Postal code"]
        ])

    buffer = io.BytesIO()
    wb_saida.save(buffer)
    buffer.seek(0)

    await update.message.reply_document(document=buffer, filename=nome_final)
    await update.message.reply_text("✅ Planilha atualizada! Boa rota! 🚀")

async def mensagem_invalida(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📄 Só consigo trabalhar com arquivos `.xlsx` no momento.\nEnvie sua planilha de rota que eu organizo pra você! 🚚")

app_telegram.add_handler(MessageHandler(filters.Document.ALL, tratar_arquivo))
app_telegram.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, mensagem_invalida))

@app_web.route(f"/{TOKEN}", methods=["POST"])
def webhook():
    update = Update.de_json(request.get_json(force=True), app_telegram.bot)
    app_telegram.update.update(update)
    return "OK", 200

@app_web.route("/")
def index():
    return "✅ Bot Rota Certa está online e aguardando planilhas!"

async def iniciar_bot():
    await app_telegram.bot.set_webhook(f"https://rota-2zrg.onrender.com/{TOKEN}")
    app_web.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))

if __name__ == "__main__":
    asyncio.run(iniciar_bot())
